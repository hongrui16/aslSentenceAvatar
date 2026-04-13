import re
from openpyxl import load_workbook
from collections import Counter

xlsx_path = "how2sign_train.xlsx"  # change to your path

wb = load_workbook(xlsx_path, read_only=True)
ws = wb.active

sentences = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[6]:
        sentences.append(str(row[6]))

word_counter = Counter()
for sent in sentences:
    words = re.findall(r"\b[a-z]+\b", sent.lower())
    word_counter.update(words)

print(f"Total sentences:    {len(sentences):,}")
print(f"Total word tokens:  {sum(word_counter.values()):,}")
print(f"Unique words:       {len(word_counter):,}")
print(f"\nTop 20 most frequent words:")
for word, count in word_counter.most_common(20):
    print(f"  {word}: {count:,}")